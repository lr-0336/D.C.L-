'''
----------------------------------exel解密---------------------------------------
                                                                                '''
import decry
from openpyxl import load_workbook

def unexel(name):

    workbook = load_workbook(filename = name)

    names = (workbook.sheetnames)

    for sheet in names:

        rows = list(range(1,workbook[sheet].max_row+1))

        columns = list(range(1,workbook[sheet].max_column+1))

        for r in rows:

            for c in columns:

                words = workbook[sheet].cell(row=r,column=c).value

                if words == None:

                    continue

                else:
                    print(words)
                    
                    workbook[sheet].cell(row=r,column=c).value = decry.unlock(words)
                    print(workbook[sheet].cell(row=r,column=c).value)
    workbook.save(filename = name)

    return 0


if __name__ == '__main__': #测试

    exelun('C:\\Users\\恩\\Desktop\\班\\计科一班.xlsx')
