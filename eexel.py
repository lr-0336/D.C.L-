'''
----------------------------------exel加密---------------------------------------
                                                                                '''
import encry

from openpyxl import load_workbook

def exelo(name):


    workbook = load_workbook(filename = name)

    names = (workbook.sheetnames)

    print(names)

    for sheet in names:

        print(sheet)

        rows = list(range(1,workbook[sheet].max_row+1))

        columns = list(range(1,workbook[sheet].max_column+1))

        for r in rows:

            for c in columns:

                words = workbook[sheet].cell(row=r,column=c).value

                if words == None:

                    continue

                else:
                    print(words)
                        
                    workbook[sheet].cell(row=r,column=c).value = encry.lock_up(str(words).replace( ' ' , '' ))

                    print(workbook[sheet].cell(row=r,column=c).value)

    workbook.save(filename = name)

    return 0



if __name__ == '__main__': #测试

    print(exelo(''))
